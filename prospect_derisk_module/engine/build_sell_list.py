#!/usr/bin/env python3
"""
build_sell_list.py  — WS2 specific-lot sell list per tax tier
Finesse Funds · Prospect Portfolio Calculator

Closes the gap between DeRisk_Tiers (which shows only per-tier TOTALS) and
Decision_Lots (which shows every lot but no tier). For each tax budget tier we
greedily select the lots that remove the most market exposure per tax dollar
(loss/zero-tax harvest lots first), then TAG every selected lot with the cheapest
tier at which it gets sold. Output:
  - sell_list.json
  - a "Sell_List" sheet appended to Portfolio_Proposal_Workbook.xlsx
    * Tier menu (recap)
    * Position-level rollup per tier (digestible for the prospect)
    * Full lot-level list, each lot tagged with its entry tier

Selection logic is identical to build_sell_tiers.py so totals reconcile exactly.
"""
import json
from pathlib import Path
from collections import defaultdict

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

BASE = Path(__file__).parent
da   = json.load(open(BASE / "decision_analysis.json"))
corr = json.load(open(BASE / "corrected_portfolio.json"))["summary"]
lots = da["lots"]

TOTAL   = corr["grand_total_current"]
CASH0   = corr["cash_mv"]
EQUITY0 = TOTAL - CASH0
DIST    = 0.05 * TOTAL
BUDGETS = [250_000, 500_000, 750_000]
DD      = [20, 30, 40]

BETA_D_TOTAL = sum(l["market_value"] * l["stress_beta"] for l in lots)
BETA_INCL_CASH = BETA_D_TOTAL / TOTAL

# rank lots: zero/negative-tax first, then beta-$ per tax-$ desc  (== build_sell_tiers.py)
def eff(l):
    t = l["tax_to_sell_marginal"]
    bd = l["market_value"] * l["stress_beta"]
    if t <= 0:
        return (0, -bd)
    return (1, -(bd / t))
ranked = sorted(lots, key=eff)


def select(budget):
    """Return the list of selected lots for a tax budget (same greedy as build_sell_tiers)."""
    sel = []
    tax = 0.0
    for l in ranked:
        t = max(l["tax_to_sell_marginal"], 0)
        if tax + t > budget:
            continue
        sel.append(l)
        tax += t
    return sel


# ---- selections per tier + tier tagging -----------------------------------
def lot_key(l):
    return (l["ticker"], l.get("trade_date"), l.get("holding_period"),
            round(l["market_value"], 2), round(l.get("unrealized_gl", 0), 2))

selections = {b: select(b) for b in BUDGETS}

# cheapest tier at which each lot is sold
tier_of = {}
for b in BUDGETS:                      # ascending budgets
    for l in selections[b]:
        k = lot_key(l)
        if k not in tier_of:
            tier_of[k] = b

# per-tier aggregates (reconcile against sell_tiers.json)
def summarize(sel):
    proceeds = sum(l["market_value"] for l in sel)
    gross_tax = sum(max(l["tax_to_sell_marginal"], 0) for l in sel)
    net_tax  = sum(l["tax_to_sell_marginal"] for l in sel)
    beta_d_removed = sum(l["market_value"] * l["stress_beta"] for l in sel)
    new_total = TOTAL - gross_tax
    new_cash  = CASH0 + (proceeds - gross_tax)
    new_beta  = (BETA_D_TOTAL - beta_d_removed) / new_total
    return {
        "n_lots": len(sel), "proceeds": proceeds, "gross_tax": gross_tax,
        "net_tax_incl_harvest": net_tax, "beta_dollars_removed": beta_d_removed,
        "new_total": new_total, "new_cash": new_cash,
        "new_cash_pct": new_cash / new_total * 100,
        "beta_after": new_beta, "runway_after": new_cash / DIST,
        "protection": {f"dd_{d}": beta_d_removed * d / 100 for d in DD},
    }

tier_summ = {b: summarize(selections[b]) for b in BUDGETS}

# ---- assemble tagged lot list (only lots that are ever sold) --------------
sold_lots = []
for l in ranked:
    k = lot_key(l)
    if k in tier_of:
        bd = l["market_value"] * l["stress_beta"]
        t = l["tax_to_sell_marginal"]
        sold_lots.append({
            "ticker": l["ticker"], "name": l.get("name", ""),
            "trade_date": l.get("trade_date"), "term": l.get("holding_period"),
            "quantity": l.get("quantity"), "market_value": l["market_value"],
            "unrealized_gl": l.get("unrealized_gl"), "stress_beta": l["stress_beta"],
            "tax_to_sell": t, "beta_dollars_removed": bd,
            "exposure_per_tax_dollar": (bd / t) if t > 0 else None,  # None => harvest/free
            "entry_tier": tier_of[k],
        })

# position-level rollup per tier (incremental — lots whose entry_tier == this tier)
def position_rollup(entry_tier):
    agg = defaultdict(lambda: {"name": "", "term_lt": 0, "lots": 0, "qty": 0.0,
                               "mv": 0.0, "gl": 0.0, "tax": 0.0, "betad": 0.0, "beta": 0.0})
    for r in sold_lots:
        if r["entry_tier"] != entry_tier:
            continue
        a = agg[r["ticker"]]
        a["name"] = r["name"]; a["beta"] = r["stress_beta"]
        a["lots"] += 1; a["qty"] += r["quantity"] or 0
        a["mv"] += r["market_value"]; a["gl"] += r["unrealized_gl"] or 0
        a["tax"] += r["tax_to_sell"]; a["betad"] += r["beta_dollars_removed"]
    rows = []
    for tk, a in agg.items():
        rows.append({"ticker": tk, **a})
    rows.sort(key=lambda x: -x["betad"])
    return rows

incremental = {b: position_rollup(b) for b in BUDGETS}

json.dump({
    "hold_all": {"total": TOTAL, "cash": CASH0, "beta_incl_cash": BETA_INCL_CASH,
                 "distribution_5pct": DIST, "runway_years": CASH0 / DIST},
    "tier_summary": tier_summ,
    "sold_lots": sold_lots,
    "incremental_positions": incremental,
}, open(BASE / "sell_list.json", "w"), indent=1, default=str)

# ===========================================================================
#  Workbook sheet
# ===========================================================================
WB = BASE / "Portfolio_Proposal_Workbook.xlsx"
wb = openpyxl.load_workbook(WB)
if "Sell_List" in wb.sheetnames:
    del wb["Sell_List"]
ws = wb.create_sheet("Sell_List")

NAVY = "1F3864"; BLUE = "2E5496"; LBLUE = "D6E0F0"; GREEN = "E2EFDA"; GREY = "F2F2F2"
thin = Side(style="thin", color="BFBFBF")
border = Border(left=thin, right=thin, top=thin, bottom=thin)
def cell(r, c, v, bold=False, fill=None, color=None, fmt=None, align="left", size=10, box=False):
    x = ws.cell(row=r, column=c, value=v)
    x.font = Font(bold=bold, color=color or "000000", size=size)
    if fill: x.fill = PatternFill("solid", fgColor=fill)
    if fmt: x.number_format = fmt
    x.alignment = Alignment(horizontal=align, vertical="center", wrap_text=False)
    if box: x.border = border
    return x

USD = '#,##0'; USD2 = '#,##0.00'; PCT = '0.0%'; NUM2 = '0.00'

r = 1
cell(r, 1, "Specific Sell List by Tax Tier", bold=True, color="FFFFFF", fill=NAVY, size=14)
for c in range(2, 9): cell(r, c, "", fill=NAVY)
r += 1
cell(r, 1, "For each tax budget the prospect chooses, these are the exact lots to sell — ranked by market exposure removed per tax dollar (loss/free lots first). Tagged by the cheapest tier each lot enters.", size=9, color="595959")
r += 2

# --- Tier menu recap -------------------------------------------------------
cell(r, 1, "THE MENU — pick your tax appetite", bold=True, color="FFFFFF", fill=BLUE, size=11)
for c in range(2, 9): cell(r, c, "", fill=BLUE)
r += 1
hdr = ["Option", "Lots sold", "Proceeds to cash", "Tax paid", "Beta after",
       "Cash % after", "Runway (yrs)", "Protection @-30%"]
for c, h in enumerate(hdr, 1):
    cell(r, c, h, bold=True, fill=LBLUE, box=True, align="center")
r += 1
cell(r, 1, "Hold all (do nothing)", box=True)
cell(r, 2, 0, box=True, align="center")
cell(r, 3, 0, fmt=USD, box=True, align="right")
cell(r, 4, 0, fmt=USD, box=True, align="right")
cell(r, 5, round(BETA_INCL_CASH, 3), fmt=NUM2, box=True, align="center")
cell(r, 6, CASH0 / TOTAL, fmt=PCT, box=True, align="center")
cell(r, 7, round(CASH0 / DIST, 1), fmt=NUM2, box=True, align="center")
cell(r, 8, 0, fmt=USD, box=True, align="right")
r += 1
for b in BUDGETS:
    s = tier_summ[b]
    cell(r, 1, f"Spend up to ${b:,.0f} tax", bold=True, box=True, fill=GREEN)
    cell(r, 2, s["n_lots"], box=True, align="center", fill=GREEN)
    cell(r, 3, round(s["proceeds"]), fmt=USD, box=True, align="right", fill=GREEN)
    cell(r, 4, round(s["gross_tax"]), fmt=USD, box=True, align="right", fill=GREEN)
    cell(r, 5, round(s["beta_after"], 3), fmt=NUM2, box=True, align="center", fill=GREEN)
    cell(r, 6, s["new_cash_pct"] / 100, fmt=PCT, box=True, align="center", fill=GREEN)
    cell(r, 7, round(s["runway_after"], 1), fmt=NUM2, box=True, align="center", fill=GREEN)
    cell(r, 8, round(s["protection"]["dd_30"]), fmt=USD, box=True, align="right", fill=GREEN)
    r += 1
r += 1

# --- Position-level incremental rollup per tier ----------------------------
cell(r, 1, "WHAT TO SELL — by position, shown as each tier ADDS on top of the prior one",
     bold=True, color="FFFFFF", fill=BLUE, size=11)
for c in range(2, 9): cell(r, c, "", fill=BLUE)
r += 1
cell(r, 1, "Tiers are cumulative: the $500K plan = the $250K sells PLUS the rows below it, and so on.", size=9, color="595959")
r += 1

cols = ["Position", "Lots", "Shares", "Market value", "Unrealized gain",
        "Tax to sell", "Beta-$ removed", "Exposure per $ tax"]
for b in BUDGETS:
    cell(r, 1, f"TIER ADD — up to ${b:,.0f} tax", bold=True, color="FFFFFF", fill=NAVY)
    for c in range(2, 9): cell(r, c, "", fill=NAVY)
    r += 1
    for c, h in enumerate(cols, 1):
        cell(r, c, h, bold=True, fill=LBLUE, box=True, align="center")
    r += 1
    rows = incremental[b]
    if not rows:
        cell(r, 1, "(no incremental positions)", box=True); r += 1
    for row in rows:
        eptd = (row["betad"] / row["tax"]) if row["tax"] > 0 else None
        cell(r, 1, f"{row['ticker']} — {row['name']}"[:42], box=True)
        cell(r, 2, row["lots"], box=True, align="center")
        cell(r, 3, round(row["qty"], 2), fmt=USD2, box=True, align="right")
        cell(r, 4, round(row["mv"]), fmt=USD, box=True, align="right")
        cell(r, 5, round(row["gl"]), fmt=USD, box=True, align="right")
        cell(r, 6, round(row["tax"]), fmt=USD, box=True, align="right")
        cell(r, 7, round(row["betad"]), fmt=USD, box=True, align="right")
        cell(r, 8, "harvest/free" if eptd is None else round(eptd, 1),
             fmt=None if eptd is None else NUM2, box=True, align="right")
        r += 1
    s = tier_summ[b]
    cell(r, 1, f"Subtotal at ${b:,.0f} tier (cumulative)", bold=True, fill=GREY, box=True)
    cell(r, 2, s["n_lots"], bold=True, fill=GREY, box=True, align="center")
    cell(r, 4, round(s["proceeds"]), bold=True, fmt=USD, fill=GREY, box=True, align="right")
    cell(r, 6, round(s["gross_tax"]), bold=True, fmt=USD, fill=GREY, box=True, align="right")
    cell(r, 7, round(s["beta_dollars_removed"]), bold=True, fmt=USD, fill=GREY, box=True, align="right")
    cell(r, 8, f"beta {s['beta_after']:.2f}", bold=True, fill=GREY, box=True, align="right")
    r += 2

# --- Full lot-level list ---------------------------------------------------
cell(r, 1, "FULL LOT DETAIL (every lot sold, tagged with its entry tier)",
     bold=True, color="FFFFFF", fill=BLUE, size=11)
for c in range(2, 10): cell(r, c, "", fill=BLUE)
r += 1
lcols = ["Entry tier", "Ticker", "Trade date", "Term", "Shares", "Market value",
         "Unrealized gain", "Stress beta", "Tax to sell", "Beta-$ removed"]
for c, h in enumerate(lcols, 1):
    cell(r, c, h, bold=True, fill=LBLUE, box=True, align="center")
r += 1
# order: by entry tier, then efficiency (already in `ranked` order within sold_lots)
order = {b: i for i, b in enumerate(BUDGETS)}
for row in sorted(sold_lots, key=lambda x: (order[x["entry_tier"]],)):
    cell(r, 1, f"${row['entry_tier']:,.0f}", box=True, align="center")
    cell(r, 2, row["ticker"], box=True)
    cell(r, 3, str(row["trade_date"]), box=True, align="center")
    cell(r, 4, row["term"], box=True, align="center")
    cell(r, 5, round(row["quantity"], 2) if row["quantity"] else "", fmt=USD2, box=True, align="right")
    cell(r, 6, round(row["market_value"]), fmt=USD, box=True, align="right")
    cell(r, 7, round(row["unrealized_gl"] or 0), fmt=USD, box=True, align="right")
    cell(r, 8, round(row["stress_beta"], 3), fmt=NUM2, box=True, align="center")
    cell(r, 9, round(row["tax_to_sell"]), fmt=USD, box=True, align="right")
    cell(r, 10, round(row["beta_dollars_removed"]), fmt=USD, box=True, align="right")
    r += 1

# widths
widths = {1: 30, 2: 14, 3: 16, 4: 14, 5: 14, 6: 16, 7: 16, 8: 14, 9: 14, 10: 16}
for col, w in widths.items():
    ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = w
ws.freeze_panes = "A4"

wb.save(WB)
print("Saved Sell_List sheet. Sold lots:", len(sold_lots))
for b in BUDGETS:
    s = tier_summ[b]
    print(f"  ${b:,} tier: {s['n_lots']} lots | tax ${s['gross_tax']:,.0f} | "
          f"proceeds ${s['proceeds']:,.0f} | beta {BETA_INCL_CASH:.2f}->{s['beta_after']:.2f} | "
          f"incremental positions {len(incremental[b])}")
